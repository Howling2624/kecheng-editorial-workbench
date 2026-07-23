from pathlib import Path
import csv
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


script_dir = Path(__file__).resolve().parent
base_dir = script_dir.parent if script_dir.name == "scripts" else script_dir
root = base_dir / "稿件表数据" / "稿件表备份"
report_dir = base_dir / "稿件表数据" / "整合探查报告"
source_mapping_path = report_dir / "06_SQLite抽取映射_待确认.csv"


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm(value):
    value = clean(value).lower()
    return re.sub(r"[\s\u3000:：;；,，.。/\\|\-—_()（）\[\]【】]+", "", value)


def cell_ref(col_idx, label):
    if not col_idx:
        return ""
    return f"{get_column_letter(col_idx)}:{label}"


def detect_header_row(ws):
    terms = [
        "receiveddate",
        "reviseddate",
        "accepteddate",
        "publisheddate",
        "declineddate",
        "publisheddeclinedate",
        "文章标题",
        "稿件类别",
        "第一作者",
        "作者信息",
        "h指数",
        "机构",
        "国家",
        "当前状态",
        "稿件状态",
        "状态",
        "round1reviewing",
        "ojs系统编号",
        "系统编号",
        "稿件编号",
        "稿件号",
        "number编号",
    ]
    best = (0, 1)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 30, 30), values_only=True), start=1):
        values = [clean(v) for v in row]
        nonempty = [v for v in values if v]
        if len(nonempty) < 5:
            continue
        joined = " ".join(norm(v) for v in nonempty)
        score = sum(1 for term in terms if term in joined)
        score += min(len(nonempty), 120) / 100
        if score > best[0]:
            best = (score, row_idx)
    return best[1]


def read_headers(ws, header_row):
    values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = []
    for idx, value in enumerate(values, start=1):
        label = clean(value)
        if label:
            headers.append({"col": idx, "label": label, "norm": norm(label)})
    return headers


def find_first(headers, predicates):
    for predicate in predicates:
        for header in headers:
            if predicate(header["label"], header["norm"]):
                return header
    return None


def find_id(headers):
    return find_first(
        headers,
        [
            lambda label, n: "稿件编号" in label,
            lambda label, n: "稿件号" in label,
            lambda label, n: "ojs系统编号" in n,
            lambda label, n: n == "系统编号",
            lambda label, n: "number编号" in n,
            lambda label, n: label == "序号",
            lambda label, n: "编号" in label,
        ],
    )


def find_status(headers):
    return find_first(
        headers,
        [
            lambda label, n: label == "当前状态",
            lambda label, n: label == "稿件状态",
            lambda label, n: label == "状态",
            lambda label, n: n == "status",
            lambda label, n: "状态" in label and "跟进" not in label,
        ],
    )


def find_date(headers, kind):
    if kind == "received":
        return find_first(headers, [lambda label, n: "receiveddate" in n])
    if kind == "revised":
        return find_first(headers, [lambda label, n: "reviseddate" in n])
    if kind == "accepted":
        return find_first(headers, [lambda label, n: "accepteddate" in n])
    if kind == "published":
        return find_first(
            headers,
            [
                lambda label, n: "publisheddate" in n and "decline" not in n,
                lambda label, n: "published" in n and "decline" in n,
            ],
        )
    if kind == "declined":
        return find_first(
            headers,
            [
                lambda label, n: "declineddate" in n and "published" not in n,
                lambda label, n: "declined" in n and "published" not in n,
                lambda label, n: "decline" in n and "published" in n,
            ],
        )
    return None


def date_rule(header, kind):
    if not header:
        return "未找到"
    n = header["norm"]
    if "published" in n and "decline" in n:
        if kind == "published":
            return "与Declined共用一列；当前状态为Published时取此列"
        if kind == "declined":
            return "与Published共用一列；当前状态为Rejected/Declined时取此列"
    return "直接取此列"


def is_author_name_header(label, normalized):
    if label in {"第一作者", "作者信息", "最高H Co-Author", "通讯作者", "通讯信息"}:
        return True
    if re.search(r"Co-?Author\s*\d+", label, re.I):
        return True
    if re.search(r"Co-?Author\d+", label, re.I):
        return True
    if re.search(r"CoAuthor\d+", normalized, re.I):
        return True
    return False


def author_role(label):
    if label in {"第一作者", "作者信息"}:
        return "first_author"
    if label == "最高H Co-Author":
        return "highest_h_coauthor"
    if "通讯" in label:
        return "corresponding_author"
    match = re.search(r"(\d+)", label)
    if match:
        return f"coauthor_{match.group(1)}"
    return "author"


def parse_author_groups(headers):
    author_headers = [h for h in headers if is_author_name_header(h["label"], h["norm"])]
    groups = []
    for idx, author in enumerate(author_headers):
        next_col = author_headers[idx + 1]["col"] if idx + 1 < len(author_headers) else 10**9
        window = [h for h in headers if author["col"] < h["col"] < next_col and h["col"] <= author["col"] + 5]
        h_index = find_first(window, [lambda label, n: label == "H指数" or n in {"h指数", "hindex"}])
        institution = find_first(window, [lambda label, n: label == "机构" or label == "单位"])
        country = find_first(window, [lambda label, n: label == "国家"])
        groups.append(
            {
                "role": author_role(author["label"]),
                "name_col": cell_ref(author["col"], author["label"]),
                "h_index_col": cell_ref(h_index["col"], h_index["label"]) if h_index else "",
                "institution_col": cell_ref(institution["col"], institution["label"]) if institution else "",
                "country_col": cell_ref(country["col"], country["label"]) if country else "",
            }
        )
    return groups


def brief_author_mapping(groups):
    parts = []
    for group in groups:
        bits = [group["role"], f"name={group['name_col']}"]
        if group["h_index_col"]:
            bits.append(f"h={group['h_index_col']}")
        if group["institution_col"]:
            bits.append(f"inst={group['institution_col']}")
        if group["country_col"]:
            bits.append(f"country={group['country_col']}")
        parts.append("(" + ", ".join(bits) + ")")
    return "; ".join(parts)


def write_csv(path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


with source_mapping_path.open(encoding="utf-8-sig") as file:
    targets = list(csv.DictReader(file))

main_rows = []
author_rows = []

for target in targets:
    journal = target["journal_code"]
    file_name = target["source_file"]
    sheet_name = target["source_sheet"]
    path = root / file_name
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header_row = detect_header_row(ws)
    headers = read_headers(ws, header_row)

    id_header = find_id(headers)
    received_header = find_date(headers, "received")
    revised_header = find_date(headers, "revised")
    accepted_header = find_date(headers, "accepted")
    published_header = find_date(headers, "published")
    declined_header = find_date(headers, "declined")
    title_header = find_first(headers, [lambda label, n: label == "文章标题", lambda label, n: "文章标题" in label])
    category_header = find_first(headers, [lambda label, n: label == "稿件类别", lambda label, n: "稿件类别" in label])
    status_header = find_status(headers)
    round1_header = find_first(headers, [lambda label, n: "round1reviewing" in n])
    author_groups = parse_author_groups(headers)

    notes = []
    if id_header and id_header["label"] == "序号":
        notes.append("稿件编号源列为“序号”，建议确认该列确为稿件编号")
    if published_header and declined_header and published_header["col"] == declined_header["col"]:
        notes.append("Published Date和Declined Date共用同一列，需按当前状态拆分")
    if not title_header:
        notes.append("未找到文章标题列")
    if not status_header:
        notes.append("未找到当前状态列")
    if not author_groups:
        notes.append("未识别到作者列组")

    main_rows.append(
        {
            "journal_code": journal,
            "source_file": file_name,
            "source_sheet": sheet_name,
            "detected_header_row": header_row,
            "manuscript_id_raw": cell_ref(id_header["col"], id_header["label"]) if id_header else "",
            "manuscript_id_clean_rule": "保留原值；清洗值去空格/换行，数字去.0，取首个编号型片段，去掉括号或空格后的备注",
            "received_date": cell_ref(received_header["col"], received_header["label"]) if received_header else "",
            "revised_date": cell_ref(revised_header["col"], revised_header["label"]) if revised_header else "",
            "accepted_date": cell_ref(accepted_header["col"], accepted_header["label"]) if accepted_header else "",
            "published_date": cell_ref(published_header["col"], published_header["label"]) if published_header else "",
            "published_date_rule": date_rule(published_header, "published"),
            "declined_date": cell_ref(declined_header["col"], declined_header["label"]) if declined_header else "",
            "declined_date_rule": date_rule(declined_header, "declined"),
            "article_title": cell_ref(title_header["col"], title_header["label"]) if title_header else "",
            "manuscript_category": cell_ref(category_header["col"], category_header["label"]) if category_header else "",
            "current_status": cell_ref(status_header["col"], status_header["label"]) if status_header else "",
            "round1_reviewing_date": cell_ref(round1_header["col"], round1_header["label"]) if round1_header else "",
            "author_group_count": len(author_groups),
            "author_groups_brief": brief_author_mapping(author_groups),
            "notes": "；".join(notes),
        }
    )

    for order, group in enumerate(author_groups, start=1):
        author_rows.append(
            {
                "journal_code": journal,
                "source_file": file_name,
                "source_sheet": sheet_name,
                "detected_header_row": header_row,
                "author_order": order,
                "author_role": group["role"],
                "author_name": group["name_col"],
                "h_index": group["h_index_col"],
                "institution": group["institution_col"],
                "country": group["country_col"],
            }
        )
    wb.close()

main_path = report_dir / "08_稿件主表字段映射_待确认.csv"
authors_path = report_dir / "09_作者字段映射_待确认.csv"
md_path = report_dir / "08_稿件字段抽取映射报告_待确认.md"

write_csv(main_path, main_rows, list(main_rows[0].keys()))
write_csv(authors_path, author_rows, list(author_rows[0].keys()))

lines = [
    "# 稿件字段抽取映射报告（待确认）",
    "",
    "## 拟采用的整理方式",
    "",
    "- `manuscripts`：一行一篇稿件，保留稿件编号、日期、标题、类别、状态、Round 1 Reviewing等字段。",
    "- `manuscript_authors`：一行一位作者，通过 `journal_code + manuscript_id_clean` 关联回稿件主表。",
    "- 稿件编号同时保留原始值和清洗值；去重优先用 `journal_code + manuscript_id_clean`。",
    "- Published/Declined 共用一列时，先按当前状态拆分：Published取为Published Date，Rejected/Declined取为Declined Date。",
    "",
    "## 稿件主表字段映射",
    "",
    "| 期刊 | Sheet | 表头行 | 稿件编号 | Received | Revised | Accepted | Published | Declined | 标题 | 类别 | 状态 | Round 1 | 备注 |",
    "|---|---|---:|---|---|---|---|---|---|---|---|---|---|---|",
]
for row in main_rows:
    lines.append(
        f"| {row['journal_code']} | {row['source_sheet']} | {row['detected_header_row']} | "
        f"{row['manuscript_id_raw']} | {row['received_date']} | {row['revised_date']} | "
        f"{row['accepted_date']} | {row['published_date']} | {row['declined_date']} | "
        f"{row['article_title']} | {row['manuscript_category']} | {row['current_status']} | "
        f"{row['round1_reviewing_date']} | {row['notes']} |"
    )

lines.extend(["", "## 作者字段映射（简表）", ""])
for row in main_rows:
    lines.append(f"### {row['journal_code']} - {row['source_sheet']}")
    lines.append("")
    lines.append(row["author_groups_brief"] or "未识别到作者字段")
    lines.append("")

md_path.write_text("\n".join(lines), encoding="utf-8")

print(main_path)
print(authors_path)
print(md_path)
