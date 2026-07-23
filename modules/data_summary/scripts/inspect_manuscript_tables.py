from pathlib import Path
from datetime import datetime
import csv
import json
import re
import hashlib
from collections import Counter, defaultdict

from openpyxl import load_workbook


script_dir = Path(__file__).resolve().parent
base_dir = script_dir.parent if script_dir.name == "scripts" else script_dir
root = base_dir / "稿件表数据" / "稿件表备份"
out_dir = base_dir / "稿件表数据" / "整合探查报告"
out_dir.mkdir(parents=True, exist_ok=True)

supported = {".xlsx", ".xlsm", ".xltx", ".xltm", ".csv", ".tsv", ".xls"}
files = sorted([p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in supported])

file_rows = []
sheet_rows = []
field_rows = []
errors = []
header_signatures = defaultdict(list)
field_counter = Counter()
normalized_field_counter = Counter()
journal_counter = Counter()


def clean_cell(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    return str(value).strip()


def norm_header(value):
    value = clean_cell(value).lower()
    return re.sub(r"[\s\u3000:：;；,，.。/\\|\-—_()（）\[\]【】]+", "", value)


def guess_journal(path):
    name = path.stem
    name = re.sub(r"(稿件表|备份|数据|汇总|副本|copy).*$", "", name, flags=re.I).strip(" _-—")
    return name or path.stem


def classify_sheet(sheet_name):
    name = sheet_name.strip().lower()
    if "下拉列表" in sheet_name:
        return "lookup_list"
    if any(keyword in sheet_name for keyword in ["稿件信息", "稿件总情况"]):
        return "manuscript_core"
    if any(keyword in sheet_name for keyword in ["稿件进度", "稿件跟进", "历史资料", "已出版稿件", "收钱的稿件", "出刊", "发刊", "意向投稿"]):
        return "manuscript_related"
    if any(keyword.lower() in name for keyword in ["审稿人", "编委", "ebm", "ae", "abm", "eab", "specialist", "顾问"]):
        return "people_or_editorial_board"
    if any(keyword in sheet_name for keyword in ["不用再发", "下架", "大哥们"]):
        return "exclusion_or_status_list"
    if any(keyword in sheet_name for keyword in ["共作", "会议合作", "关键词", "特殊注意事项"]):
        return "reference_or_operations"
    return "other"


def sha1_short(path):
    digest = hashlib.sha1()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()[:12]


def detect_header(rows):
    scored = []
    for row_number, values in rows:
        cleaned = [clean_cell(value) for value in values]
        nonempty = [value for value in cleaned if value]
        if not nonempty:
            continue
        unique_nonempty = len(set(nonempty))
        avg_len = sum(len(value) for value in nonempty) / len(nonempty)
        score = len(nonempty) * 10 + unique_nonempty * 2
        if avg_len <= 30:
            score += 8
        if len(nonempty) >= 3:
            score += 8
        if len(nonempty) <= 2:
            score -= 20
        scored.append((score, row_number, cleaned, len(nonempty), unique_nonempty))
    if not scored:
        return None, [], 0
    scored.sort(key=lambda item: (item[0], item[3], item[4]), reverse=True)
    _, row_number, cleaned, nonempty_count, _ = scored[0]
    return row_number, cleaned, nonempty_count


for path in files:
    suffix = path.suffix.lower()
    journal = guess_journal(path)
    journal_counter[journal] += 1
    stat = path.stat()
    file_info = {
        "journal_guess": journal,
        "file_name": path.name,
        "relative_path": str(path.relative_to(root)),
        "full_path": str(path),
        "extension": suffix,
        "size_bytes": stat.st_size,
        "size_mb": round(stat.st_size / 1024 / 1024, 3),
        "modified_time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "sha1_12": "",
        "sheet_count": 0,
        "status": "ok",
        "error": "",
    }
    try:
        file_info["sha1_12"] = sha1_short(path)
        if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            file_info["status"] = "unsupported_for_deep_scan"
            file_info["error"] = f"暂未深度扫描 {suffix}，后续可另配读取器"
            file_rows.append(file_info)
            continue

        workbook = load_workbook(path, read_only=False, data_only=True)
        file_info["sheet_count"] = len(workbook.worksheets)
        for worksheet in workbook.worksheets:
            sheet_type = classify_sheet(worksheet.title)
            max_row = worksheet.max_row or 0
            max_col = worksheet.max_column or 0
            top_rows = []
            limit_rows = min(max_row, 30)
            limit_cols = min(max_col, 300)
            for row in worksheet.iter_rows(min_row=1, max_row=limit_rows, max_col=limit_cols, values_only=True):
                top_rows.append((len(top_rows) + 1, list(row)))

            header_row, header_values, header_nonempty = detect_header(top_rows)
            headers = []
            duplicate_headers = []
            blank_headers = 0
            data_rows = 0
            nonempty_rows_total = 0
            nonempty_cols_observed = set()
            data_cols_after_header = set()

            if header_row:
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    has_any = False
                    for col_index, value in enumerate(row, start=1):
                        if clean_cell(value):
                            has_any = True
                            nonempty_cols_observed.add(col_index)
                            if row_index > header_row:
                                data_cols_after_header.add(col_index)
                    if has_any:
                        nonempty_rows_total += 1
                        if row_index > header_row:
                            data_rows += 1

                seen = Counter()
                header_label_cols = {
                    col_index
                    for col_index, header in enumerate(header_values[:max_col], start=1)
                    if clean_cell(header)
                }
                active_cols = sorted(header_label_cols | data_cols_after_header)

                for col_index in active_cols:
                    header = header_values[col_index - 1] if col_index <= len(header_values) else ""
                    label = clean_cell(header)
                    if not label:
                        blank_headers += 1
                        label = f"__blank_col_{col_index}"
                    seen[label] += 1
                    if seen[label] > 1:
                        duplicate_headers.append(label)
                    headers.append(label)
                    field_counter[label] += 1
                    normalized_field_counter[norm_header(label)] += 1
                    field_rows.append(
                        {
                            "journal_guess": journal,
                            "file_name": path.name,
                            "sheet_name": worksheet.title,
                            "sheet_type_guess": sheet_type,
                            "header_row": header_row,
                            "column_index": col_index,
                            "field_name": label,
                            "normalized_field": norm_header(label),
                        }
                    )
            else:
                for row in worksheet.iter_rows(values_only=True):
                    if any(clean_cell(value) for value in row):
                        nonempty_rows_total += 1

            signature = " | ".join([header for header in headers if not header.startswith("__blank_col_")])
            if signature:
                header_signatures[signature].append(f"{path.name}::{worksheet.title}")

            sheet_rows.append(
                {
                    "journal_guess": journal,
                    "file_name": path.name,
                    "sheet_name": worksheet.title,
                    "sheet_type_guess": sheet_type,
                    "sheet_state": worksheet.sheet_state,
                    "max_row": max_row,
                    "max_column": max_col,
                    "nonempty_rows_total": nonempty_rows_total,
                    "header_row_guess": header_row or "",
                    "header_nonempty_cells": header_nonempty,
                    "data_rows_after_header": data_rows if header_row else "",
                    "observed_nonempty_columns": len(nonempty_cols_observed),
                    "data_columns_after_header": len(data_cols_after_header),
                    "merged_cell_ranges": len(list(worksheet.merged_cells.ranges)),
                    "blank_header_count": blank_headers,
                    "duplicate_header_count": len(set(duplicate_headers)),
                    "duplicate_headers": "; ".join(sorted(set(duplicate_headers))),
                    "field_count_from_header": len(headers),
                    "fields_joined": " | ".join(headers),
                }
            )
        workbook.close()
        file_rows.append(file_info)
    except Exception as exc:
        file_info["status"] = "error"
        file_info["error"] = repr(exc)
        errors.append({"file_name": path.name, "full_path": str(path), "error": repr(exc)})
        file_rows.append(file_info)


sig_rows = []
for index, (signature, locations) in enumerate(
    sorted(header_signatures.items(), key=lambda item: (-len(item[1]), item[0])), start=1
):
    sig_rows.append(
        {
            "signature_id": f"SIG{index:03d}",
            "sheet_count": len(locations),
            "locations": "; ".join(locations),
            "fields_joined": signature,
        }
    )


field_groups = {
    "稿件编号/稿号": ["稿件编号", "稿号", "稿件id", "文章编号", "manuscriptid", "msid", "编号"],
    "标题/题名": ["标题", "题名", "篇名", "文章题目", "稿件题目", "title"],
    "作者": ["作者", "第一作者", "通讯作者", "author"],
    "单位/机构": ["单位", "机构", "作者单位", "affiliation", "institution"],
    "投稿/收稿日期": ["投稿日期", "收稿日期", "来稿日期", "提交日期", "received", "submitted"],
    "状态/处理结果": ["状态", "稿件状态", "处理结果", "status", "录用", "退稿"],
    "栏目/类别": ["栏目", "类别", "专业", "方向", "学科", "category", "section"],
    "关键词": ["关键词", "关键字", "keyword"],
    "基金": ["基金", "fund", "资助"],
    "备注": ["备注", "说明", "note", "comment"],
}
map_rows = []
all_fields = sorted(field_counter.keys())
for standard_field, keywords in field_groups.items():
    matches = []
    for field in all_fields:
        normalized_field = norm_header(field)
        if any(norm_header(keyword) in normalized_field for keyword in keywords):
            matches.append(field)
    map_rows.append(
        {
            "standard_field_candidate": standard_field,
            "matched_source_fields": " | ".join(matches),
            "match_count": len(matches),
        }
    )


def write_csv(path, rows, fieldnames):
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


write_csv(out_dir / "01_文件清单.csv", file_rows, list(file_rows[0].keys()) if file_rows else [])
write_csv(out_dir / "02_Sheet清单.csv", sheet_rows, list(sheet_rows[0].keys()) if sheet_rows else [])
write_csv(out_dir / "03_字段清单.csv", field_rows, list(field_rows[0].keys()) if field_rows else [])
write_csv(
    out_dir / "04_表头结构分组.csv",
    sig_rows,
    list(sig_rows[0].keys()) if sig_rows else ["signature_id", "sheet_count", "locations", "fields_joined"],
)
write_csv(out_dir / "05_标准字段候选匹配.csv", map_rows, list(map_rows[0].keys()) if map_rows else [])

ok_files = [row for row in file_rows if row["status"] == "ok"]
error_files = [row for row in file_rows if row["status"] == "error"]
unsupported_files = [row for row in file_rows if row["status"] == "unsupported_for_deep_scan"]
total_sheets = len(sheet_rows)
total_data_rows = sum(int(row["data_rows_after_header"] or 0) for row in sheet_rows)
unique_raw_fields = len(field_counter)
unique_norm_fields = len([key for key in normalized_field_counter if key])
problem_sheets = [
    row for row in sheet_rows if row["blank_header_count"] or row["duplicate_header_count"] or not row["header_row_guess"]
]
large_header_offsets = [
    row for row in sheet_rows if str(row["header_row_guess"]).isdigit() and int(row["header_row_guess"]) > 1
]
identical_sigs = [row for row in sig_rows if row["sheet_count"] > 1]
sheet_type_counter = Counter(row["sheet_type_guess"] for row in sheet_rows)
manuscript_core_rows = [row for row in sheet_rows if row["sheet_type_guess"] == "manuscript_core"]
manuscript_related_rows = [row for row in sheet_rows if row["sheet_type_guess"] == "manuscript_related"]
manuscript_scope_rows = manuscript_core_rows + manuscript_related_rows
manuscript_scope_data_rows = sum(int(row["data_rows_after_header"] or 0) for row in manuscript_scope_rows)

report_lines = [
    "# 稿件表备份数据探查报告",
    "",
    f"- 扫描目录：`{root}`",
    f"- 输出目录：`{out_dir}`",
    f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    "",
    "## 一、总体情况",
    "",
    f"- 发现表格文件：{len(files)} 个",
    f"- 可深度读取文件：{len(ok_files)} 个",
    f"- 深度扫描 sheet：{total_sheets} 个",
    f"- 估算数据行数（表头后非空行）：{total_data_rows} 行",
    f"- 初步识别稿件核心/相关 sheet：{len(manuscript_scope_rows)} 个，约 {manuscript_scope_data_rows} 行",
    f"- 原始字段名数量：{unique_raw_fields} 个",
    f"- 规范化后字段名数量：{unique_norm_fields} 个",
    f"- 猜测期刊数量：{len(journal_counter)} 个",
    "",
    "Sheet 类型初步分布：",
    "",
]
for sheet_type, count in sheet_type_counter.most_common():
    report_lines.append(f"- `{sheet_type}`：{count} 个")

report_lines.extend(
    [
        "",
        "其中 `manuscript_core` 适合优先进入稿件主表；`manuscript_related` 建议先作为稿件相关附表或候选数据源保留。",
        "",
        "## 二、文件概览",
        "",
        "| 期刊猜测 | 文件 | 大小MB | Sheet数 | 修改时间 | 状态 |",
        "|---|---|---:|---:|---|---|",
    ]
)
for row in file_rows:
    report_lines.append(
        f"| {row['journal_guess']} | {row['file_name']} | {row['size_mb']} | "
        f"{row['sheet_count']} | {row['modified_time']} | {row['status']} |"
    )

report_lines.extend(
    [
        "",
        "## 三、Sheet 与字段结构",
        "",
        "| 期刊 | 文件 | Sheet | 类型猜测 | 表头行 | 数据行 | 字段数 | 空表头 | 重复表头 | 合并单元格 |",
        "|---|---|---|---|---:|---:|---:|---:|---:|---:|",
    ]
)
for row in sheet_rows:
    report_lines.append(
        f"| {row['journal_guess']} | {row['file_name']} | {row['sheet_name']} | {row['sheet_type_guess']} | "
        f"{row['header_row_guess']} | {row['data_rows_after_header']} | {row['field_count_from_header']} | "
        f"{row['blank_header_count']} | {row['duplicate_header_count']} | {row['merged_cell_ranges']} |"
    )

report_lines.extend(["", "## 四、主要发现", ""])
if not error_files and not unsupported_files:
    report_lines.append("- 所有发现的文件均为 `.xlsx`，且均可正常读取。")
else:
    if error_files:
        report_lines.append(f"- 有 {len(error_files)} 个文件读取失败，详见 `01_文件清单.csv`。")
    if unsupported_files:
        report_lines.append(f"- 有 {len(unsupported_files)} 个非 `.xlsx` 文件暂未深度扫描。")
if total_sheets == len(ok_files):
    report_lines.append("- 每个文件目前都是 1 个 sheet，后续入库逻辑可以先按“一个文件=一个期刊表”处理。")
else:
    report_lines.append("- 存在多 sheet 文件，后续入库需要保留 sheet 来源字段。")
if large_header_offsets:
    report_lines.append(f"- 有 {len(large_header_offsets)} 个 sheet 的表头不在第 1 行，导入时需要跳过说明/标题行。")
else:
    report_lines.append("- 表头基本位于第 1 行，自动导入难度较低。")
if problem_sheets:
    report_lines.append(f"- 有 {len(problem_sheets)} 个 sheet 存在空表头、重复表头或无法识别表头，入库前建议人工确认。")
else:
    report_lines.append("- 未发现空表头、重复表头或无法识别表头的问题。")
if identical_sigs:
    report_lines.append(f"- 发现 {len(identical_sigs)} 组表头结构被多个 sheet 共用，说明至少部分期刊表结构较统一。")
else:
    report_lines.append("- 各期刊表头结构差异较大，建议后续建立字段映射表，而不是硬拼字段。")

report_lines.extend(
    [
        "",
        "## 五、建议的 SQLite 整理方向",
        "",
        "建议后续建库时保留两类数据：原始可追溯数据，以及规范化后的主表。",
        "",
        "- `source_files`：文件级元数据，记录文件名、期刊、hash、修改时间、导入时间。",
        "- `source_sheets`：sheet 级元数据，记录表头行、原字段列表、数据行数、异常标记。",
        "- `raw_manuscripts`：原始宽表，保留所有原字段，同时附加 `journal`、`source_file`、`source_sheet`、`row_number`。",
        "- `manuscripts_normalized`：规范化主表，用字段映射表抽取稿号、标题、作者、日期、状态等核心字段。",
        "- `field_mapping`：字段映射表，可人工维护不同期刊字段到标准字段的对应关系。",
        "",
        "## 六、已输出明细文件",
        "",
    ]
)
for file_name in ["01_文件清单.csv", "02_Sheet清单.csv", "03_字段清单.csv", "04_表头结构分组.csv", "05_标准字段候选匹配.csv"]:
    report_lines.append(f"- `{file_name}`")

report_lines.extend(
    [
        "",
        "## 七、需要确认的问题",
        "",
        "1. 文件名开头是否就是正式期刊代码？例如 `JOURNAL-A稿件表.xlsx` 对应期刊代码 `JOURNAL-A`。",
        "2. 后续去重时，优先使用哪个字段作为稿件唯一标识：稿件编号/稿号，还是“标题+作者+投稿日期”？",
        "3. 是否需要把历史备份中的全部字段都保留进 SQLite，还是只保留标准字段和少量原始追溯字段？",
        "",
    ]
)

report_path = out_dir / "稿件表备份数据探查报告.md"
report_path.write_text("\n".join(report_lines), encoding="utf-8")

summary = {
    "root": str(root),
    "out_dir": str(out_dir),
    "files": len(files),
    "ok_files": len(ok_files),
    "sheets": total_sheets,
    "estimated_data_rows": total_data_rows,
    "unique_raw_fields": unique_raw_fields,
    "unique_normalized_fields": unique_norm_fields,
    "problem_sheet_count": len(problem_sheets),
    "report_path": str(report_path),
    "errors": errors,
}
(out_dir / "00_探查摘要.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps(summary, ensure_ascii=False, indent=2))
