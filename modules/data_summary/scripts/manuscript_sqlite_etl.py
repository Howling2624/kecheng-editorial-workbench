from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent if SCRIPT_DIR.name == "scripts" else SCRIPT_DIR
DEFAULT_BACKUP_DIR = BASE_DIR / "稿件表数据" / "稿件表备份"
DEFAULT_CONFIG_DIR = BASE_DIR / "稿件表数据" / "整合配置"
DEFAULT_REPORT_DIR = BASE_DIR / "稿件表数据" / "整合探查报告"
DEFAULT_OUTPUT_DIR = BASE_DIR / "稿件表数据" / "整合结果"
DEFAULT_SQLITE_PATH = DEFAULT_OUTPUT_DIR / "稿件数据.sqlite"


MAIN_FIELDS = [
    "manuscript_id_raw",
    "received_date",
    "revised_date",
    "accepted_date",
    "published_date",
    "declined_date",
    "article_title",
    "manuscript_category",
    "current_status",
    "round1_reviewing_date",
]


@dataclass
class SheetConfig:
    journal_code: str
    source_file: str
    source_sheet: str
    enabled: bool
    source_priority: int


@dataclass
class FieldConfig:
    journal_code: str
    source_file: str
    source_sheet: str
    field_name: str
    column_index: int | None
    source_label: str
    rule: str


@dataclass
class AuthorConfig:
    journal_code: str
    source_file: str
    source_sheet: str
    author_order: int
    author_role: str
    author_name_col: int | None
    h_index_col: int | None
    institution_col: int | None
    country_col: int | None
    author_name_label: str
    h_index_label: str
    institution_label: str
    country_label: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def normalize_empty(value: Any) -> str:
    text = clean_text(value)
    if text in {"", "/", "\\", "-", "--", "—", "nan", "None", "#VALUE!", "#N/A"}:
        return ""
    return text


def normalize_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_empty(value)
    if not text:
        return ""
    text = text.replace("年", "-").replace("月", "-").replace("日", "")
    text = text.replace("/", "-").replace(".", "-")
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        try:
            return date(int(year), int(month), int(day)).isoformat()
        except ValueError:
            return ""
    return text


def clean_manuscript_id(value: Any) -> str:
    text = normalize_empty(value)
    if not text:
        return ""
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"[\(\（\[\【].*$", "", text).strip()
    match = re.search(r"[A-Za-z]*\d+(?:[-_][A-Za-z0-9]+)*", text)
    if match:
        return match.group(0)
    return re.split(r"[\s,，;；/]+", text)[0].strip()


def status_for_split(value: Any) -> str:
    return clean_text(value).lower()


def split_shared_pub_declined(shared_date: str, status: str, target: str) -> str:
    status_norm = status_for_split(status)
    if target == "published" and "published" in status_norm:
        return shared_date
    if target == "declined" and ("rejected" in status_norm or "declined" in status_norm):
        return shared_date
    return ""


def apply_terminal_status_date_rule(values: dict[str, str]) -> None:
    status_norm = status_for_split(values.get("current_status", ""))
    if "rejected" in status_norm or "declined" in status_norm:
        if not values.get("declined_date") and values.get("published_date"):
            values["declined_date"] = values["published_date"]
        values["published_date"] = ""
    elif "published" in status_norm:
        if not values.get("published_date") and values.get("declined_date"):
            values["published_date"] = values["declined_date"]
        values["declined_date"] = ""


def sha1_short(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()[:12]


def read_csv(path: Path) -> list[dict[str, str]]:
    last_error: Exception | None = None
    for encoding in ["utf-8-sig", "utf-16", "gb18030"]:
        try:
            with path.open(encoding=encoding, newline="") as file:
                return list(csv.DictReader(file))
        except UnicodeError as exc:
            last_error = exc
    raise UnicodeError(f"无法识别CSV编码：{path}") from last_error


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def parse_cell_ref(value: str) -> tuple[int | None, str]:
    value = clean_text(value)
    if not value:
        return None, ""
    if ":" not in value:
        return None, value
    col, label = value.split(":", 1)
    try:
        return column_index_from_string(col), label
    except ValueError:
        return None, label


def init_config(report_dir: Path, config_dir: Path) -> None:
    sheet_source = report_dir / "06_SQLite抽取映射_待确认.csv"
    main_source = report_dir / "08_稿件主表字段映射_待确认.csv"
    author_source = report_dir / "09_作者字段映射_待确认.csv"
    missing = [str(path) for path in [sheet_source, main_source, author_source] if not path.exists()]
    if missing:
        raise FileNotFoundError("缺少映射来源文件：" + "；".join(missing))

    sheet_rows = []
    for row in read_csv(sheet_source):
        journal = row["journal_code"]
        sheet = row["source_sheet"]
        try:
            priority = int(clean_text(row.get("source_priority")) or "1")
        except ValueError:
            priority = 1
        sheet_rows.append(
            {
                "journal_code": journal,
                "source_file": row["source_file"],
                "source_sheet": sheet,
                "enabled": "是",
                "source_priority": priority,
                "note": row.get("note", ""),
            }
        )

    main_rows = []
    for row in read_csv(main_source):
        for field in MAIN_FIELDS:
            col, label = parse_cell_ref(row.get(field, ""))
            rule = ""
            if field == "manuscript_id_raw":
                rule = row.get("manuscript_id_clean_rule", "")
            elif field == "published_date":
                rule = row.get("published_date_rule", "")
            elif field == "declined_date":
                rule = row.get("declined_date_rule", "")
            main_rows.append(
                {
                    "journal_code": row["journal_code"],
                    "source_file": row["source_file"],
                    "source_sheet": row["source_sheet"],
                    "field_name": field,
                    "source_column": get_column_letter(col) if col else "",
                    "source_column_index": col or "",
                    "source_label": label,
                    "rule": rule,
                }
            )

    author_rows = []
    for row in read_csv(author_source):
        parsed = {}
        labels = {}
        for key in ["author_name", "h_index", "institution", "country"]:
            col, label = parse_cell_ref(row.get(key, ""))
            parsed[key] = col
            labels[key] = label
        author_rows.append(
            {
                "journal_code": row["journal_code"],
                "source_file": row["source_file"],
                "source_sheet": row["source_sheet"],
                "author_order": row["author_order"],
                "author_role": row["author_role"],
                "author_name_column": get_column_letter(parsed["author_name"]) if parsed["author_name"] else "",
                "author_name_column_index": parsed["author_name"] or "",
                "author_name_label": labels["author_name"],
                "h_index_column": get_column_letter(parsed["h_index"]) if parsed["h_index"] else "",
                "h_index_column_index": parsed["h_index"] or "",
                "h_index_label": labels["h_index"],
                "institution_column": get_column_letter(parsed["institution"]) if parsed["institution"] else "",
                "institution_column_index": parsed["institution"] or "",
                "institution_label": labels["institution"],
                "country_column": get_column_letter(parsed["country"]) if parsed["country"] else "",
                "country_column_index": parsed["country"] or "",
                "country_label": labels["country"],
            }
        )

    write_csv(
        config_dir / "sheet_mapping.csv",
        sheet_rows,
        ["journal_code", "source_file", "source_sheet", "enabled", "source_priority", "note"],
    )
    write_csv(
        config_dir / "field_mapping.csv",
        main_rows,
        [
            "journal_code",
            "source_file",
            "source_sheet",
            "field_name",
            "source_column",
            "source_column_index",
            "source_label",
            "rule",
        ],
    )
    write_csv(
        config_dir / "author_mapping.csv",
        author_rows,
        [
            "journal_code",
            "source_file",
            "source_sheet",
            "author_order",
            "author_role",
            "author_name_column",
            "author_name_column_index",
            "author_name_label",
            "h_index_column",
            "h_index_column_index",
            "h_index_label",
            "institution_column",
            "institution_column_index",
            "institution_label",
            "country_column",
            "country_column_index",
            "country_label",
        ],
    )


def load_sheet_configs(config_dir: Path) -> list[SheetConfig]:
    configs = []
    for row in read_csv(config_dir / "sheet_mapping.csv"):
        configs.append(
            SheetConfig(
                journal_code=row["journal_code"],
                source_file=row["source_file"],
                source_sheet=row["source_sheet"],
                enabled=row["enabled"].strip() in {"是", "yes", "Y", "y", "1", "true", "True"},
                source_priority=int(row["source_priority"] or "1"),
            )
        )
    return configs


def load_field_configs(config_dir: Path) -> dict[tuple[str, str, str], dict[str, FieldConfig]]:
    grouped: dict[tuple[str, str, str], dict[str, FieldConfig]] = defaultdict(dict)
    for row in read_csv(config_dir / "field_mapping.csv"):
        key = (row["journal_code"], row["source_file"], row["source_sheet"])
        col = row.get("source_column_index", "")
        grouped[key][row["field_name"]] = FieldConfig(
            journal_code=row["journal_code"],
            source_file=row["source_file"],
            source_sheet=row["source_sheet"],
            field_name=row["field_name"],
            column_index=int(col) if col else None,
            source_label=row.get("source_label", ""),
            rule=row.get("rule", ""),
        )
    return grouped


def load_author_configs(config_dir: Path) -> dict[tuple[str, str, str], list[AuthorConfig]]:
    grouped: dict[tuple[str, str, str], list[AuthorConfig]] = defaultdict(list)
    for row in read_csv(config_dir / "author_mapping.csv"):
        key = (row["journal_code"], row["source_file"], row["source_sheet"])

        def as_int(name: str) -> int | None:
            value = row.get(name, "")
            return int(value) if value else None

        grouped[key].append(
            AuthorConfig(
                journal_code=row["journal_code"],
                source_file=row["source_file"],
                source_sheet=row["source_sheet"],
                author_order=int(row["author_order"]),
                author_role=row["author_role"],
                author_name_col=as_int("author_name_column_index"),
                h_index_col=as_int("h_index_column_index"),
                institution_col=as_int("institution_column_index"),
                country_col=as_int("country_column_index"),
                author_name_label=row.get("author_name_label", ""),
                h_index_label=row.get("h_index_label", ""),
                institution_label=row.get("institution_label", ""),
                country_label=row.get("country_label", ""),
            )
        )
    for key in grouped:
        grouped[key].sort(key=lambda item: item.author_order)
    return grouped


def row_value(row: tuple[Any, ...], col: int | None) -> Any:
    if not col or col < 1 or col > len(row):
        return None
    return row[col - 1]


def extract_field(row: tuple[Any, ...], config: FieldConfig | None) -> Any:
    if not config:
        return None
    return row_value(row, config.column_index)


def setup_database(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(
        """
        CREATE TABLE source_files (
          file_id INTEGER PRIMARY KEY,
          journal_code TEXT NOT NULL,
          source_file TEXT NOT NULL,
          full_path TEXT NOT NULL,
          size_bytes INTEGER,
          modified_time TEXT,
          sha1_12 TEXT,
          imported_at TEXT NOT NULL
        );

        CREATE TABLE source_sheets (
          sheet_id INTEGER PRIMARY KEY,
          file_id INTEGER NOT NULL,
          journal_code TEXT NOT NULL,
          source_file TEXT NOT NULL,
          source_sheet TEXT NOT NULL,
          header_row INTEGER NOT NULL,
          data_rows_scanned INTEGER NOT NULL,
          data_rows_imported INTEGER NOT NULL,
          source_priority INTEGER NOT NULL,
          FOREIGN KEY(file_id) REFERENCES source_files(file_id)
        );

        CREATE TABLE raw_manuscripts (
          raw_id INTEGER PRIMARY KEY,
          journal_code TEXT NOT NULL,
          source_file TEXT NOT NULL,
          source_sheet TEXT NOT NULL,
          source_row_number INTEGER NOT NULL,
          source_priority INTEGER NOT NULL,
          manuscript_id_raw TEXT,
          manuscript_id_clean TEXT,
          row_json TEXT NOT NULL,
          imported_at TEXT NOT NULL
        );

        CREATE TABLE manuscripts (
          manuscript_key TEXT PRIMARY KEY,
          journal_code TEXT NOT NULL,
          manuscript_id_clean TEXT,
          manuscript_id_raw TEXT,
          received_date TEXT,
          revised_date TEXT,
          accepted_date TEXT,
          published_date TEXT,
          declined_date TEXT,
          article_title TEXT,
          manuscript_category TEXT,
          current_status TEXT,
          round1_reviewing_date TEXT,
          source_file TEXT NOT NULL,
          source_sheet TEXT NOT NULL,
          source_row_number INTEGER NOT NULL,
          source_priority INTEGER NOT NULL,
          duplicate_count INTEGER NOT NULL,
          imported_at TEXT NOT NULL
        );

        CREATE TABLE manuscript_authors (
          author_id INTEGER PRIMARY KEY,
          manuscript_key TEXT NOT NULL,
          journal_code TEXT NOT NULL,
          manuscript_id_clean TEXT,
          source_file TEXT NOT NULL,
          source_sheet TEXT NOT NULL,
          source_row_number INTEGER NOT NULL,
          author_order INTEGER NOT NULL,
          author_role TEXT,
          author_name TEXT,
          h_index TEXT,
          institution TEXT,
          country TEXT,
          FOREIGN KEY(manuscript_key) REFERENCES manuscripts(manuscript_key)
        );

        CREATE TABLE etl_warnings (
          warning_id INTEGER PRIMARY KEY,
          journal_code TEXT,
          source_file TEXT,
          source_sheet TEXT,
          source_row_number INTEGER,
          warning_type TEXT NOT NULL,
          message TEXT NOT NULL
        );

        CREATE INDEX idx_raw_manuscripts_key ON raw_manuscripts(journal_code, manuscript_id_clean);
        CREATE INDEX idx_manuscripts_status ON manuscripts(current_status);
        CREATE INDEX idx_manuscripts_dates ON manuscripts(received_date, accepted_date, published_date, declined_date);
        CREATE INDEX idx_authors_name ON manuscript_authors(author_name);
        """
    )
    return conn


def publish_database(staged_path: Path, target_path: Path) -> Path | None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path = None

    if target_path.exists():
        backup_path = target_path.with_name(
            target_path.stem
            + "_上次运行备份_"
            + datetime.now().strftime("%Y%m%d_%H%M%S")
            + target_path.suffix
        )
        source = sqlite3.connect(f"file:{target_path.as_posix()}?mode=ro", uri=True, timeout=30)
        backup = sqlite3.connect(backup_path)
        try:
            source.backup(backup)
        finally:
            backup.close()
            source.close()

    try:
        staged_path.replace(target_path)
        return backup_path
    except PermissionError:
        source = sqlite3.connect(staged_path, timeout=30)
        target = sqlite3.connect(target_path, timeout=30)
        try:
            source.backup(target)
            target.commit()
            try:
                target.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            except sqlite3.Error:
                pass
        finally:
            target.close()
            source.close()
        staged_path.unlink(missing_ok=True)
        return backup_path


def make_row_json(row: tuple[Any, ...], headers: list[str]) -> str:
    payload = {}
    for idx, value in enumerate(row, start=1):
        text = normalize_empty(value)
        if not text:
            continue
        label = headers[idx - 1] if idx <= len(headers) and headers[idx - 1] else get_column_letter(idx)
        payload[label] = text
    return json.dumps(payload, ensure_ascii=False, default=str)


def has_any_useful_value(values: dict[str, str]) -> bool:
    return any(values.get(name) for name in ["manuscript_id_raw", "article_title", "received_date", "current_status"])


def run_build(backup_dir: Path, config_dir: Path, sqlite_path: Path) -> dict[str, Any]:
    if not (config_dir / "sheet_mapping.csv").exists():
        init_config(DEFAULT_REPORT_DIR, config_dir)

    sheet_configs = [config for config in load_sheet_configs(config_dir) if config.enabled]
    field_configs = load_field_configs(config_dir)
    author_configs = load_author_configs(config_dir)
    working_sqlite_path = sqlite_path.with_name(
        f".{sqlite_path.stem}.building-{os.getpid()}{sqlite_path.suffix}"
    )
    conn = setup_database(working_sqlite_path)
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    candidates: dict[str, list[dict[str, Any]]] = defaultdict(list)
    candidate_authors: dict[str, list[dict[str, Any]]] = defaultdict(list)
    stats = {
        "source_sheets": 0,
        "raw_rows": 0,
        "manuscripts": 0,
        "authors": 0,
        "warnings": 0,
        "duplicates": 0,
    }

    for sheet_config in sheet_configs:
        key = (sheet_config.journal_code, sheet_config.source_file, sheet_config.source_sheet)
        field_map = field_configs.get(key, {})
        author_map = author_configs.get(key, [])
        path = backup_dir / sheet_config.source_file
        if not path.exists():
            conn.execute(
                "INSERT INTO etl_warnings (journal_code, source_file, source_sheet, warning_type, message) VALUES (?, ?, ?, ?, ?)",
                (sheet_config.journal_code, sheet_config.source_file, sheet_config.source_sheet, "missing_file", str(path)),
            )
            stats["warnings"] += 1
            continue

        stat = path.stat()
        file_cur = conn.execute(
            """
            INSERT INTO source_files
            (journal_code, source_file, full_path, size_bytes, modified_time, sha1_12, imported_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sheet_config.journal_code,
                sheet_config.source_file,
                str(path),
                stat.st_size,
                datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                sha1_short(path),
                imported_at,
            ),
        )
        file_id = file_cur.lastrowid

        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet_config.source_sheet not in workbook.sheetnames:
            conn.execute(
                "INSERT INTO etl_warnings (journal_code, source_file, source_sheet, warning_type, message) VALUES (?, ?, ?, ?, ?)",
                (
                    sheet_config.journal_code,
                    sheet_config.source_file,
                    sheet_config.source_sheet,
                    "missing_sheet",
                    "配置中的sheet不存在",
                ),
            )
            stats["warnings"] += 1
            workbook.close()
            continue
        worksheet = workbook[sheet_config.source_sheet]
        header_row = max(config.column_index or 1 for config in field_map.values())
        # The configured columns do not store header row; read it from field mapping source labels by locating the row.
        # These workbooks are stable, so a focused scan over the first 30 rows is enough.
        header_row = detect_configured_header_row(worksheet, field_map)
        header_values = next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        headers = [clean_text(value) for value in header_values]
        data_rows_scanned = 0
        data_rows_imported = 0

        sheet_cur = conn.execute(
            """
            INSERT INTO source_sheets
            (file_id, journal_code, source_file, source_sheet, header_row, data_rows_scanned, data_rows_imported, source_priority)
            VALUES (?, ?, ?, ?, ?, 0, 0, ?)
            """,
            (
                file_id,
                sheet_config.journal_code,
                sheet_config.source_file,
                sheet_config.source_sheet,
                header_row,
                sheet_config.source_priority,
            ),
        )
        sheet_id = sheet_cur.lastrowid

        for row_number, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            data_rows_scanned += 1
            values = {}
            for field in MAIN_FIELDS:
                raw_value = extract_field(row, field_map.get(field))
                if field.endswith("_date") or field == "round1_reviewing_date":
                    values[field] = normalize_date(raw_value)
                else:
                    values[field] = normalize_empty(raw_value)

            if not has_any_useful_value(values):
                continue

            values["manuscript_id_clean"] = clean_manuscript_id(values.get("manuscript_id_raw"))
            status = values.get("current_status", "")
            published_config = field_map.get("published_date")
            declined_config = field_map.get("declined_date")
            if (
                published_config
                and declined_config
                and published_config.column_index
                and published_config.column_index == declined_config.column_index
            ):
                shared_date = normalize_date(row_value(row, published_config.column_index))
                values["published_date"] = split_shared_pub_declined(shared_date, status, "published")
                values["declined_date"] = split_shared_pub_declined(shared_date, status, "declined")
            apply_terminal_status_date_rule(values)

            row_json = make_row_json(row, headers)
            conn.execute(
                """
                INSERT INTO raw_manuscripts
                (journal_code, source_file, source_sheet, source_row_number, source_priority,
                 manuscript_id_raw, manuscript_id_clean, row_json, imported_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sheet_config.journal_code,
                    sheet_config.source_file,
                    sheet_config.source_sheet,
                    row_number,
                    sheet_config.source_priority,
                    values["manuscript_id_raw"],
                    values["manuscript_id_clean"],
                    row_json,
                    imported_at,
                ),
            )
            data_rows_imported += 1
            stats["raw_rows"] += 1

            if not values["manuscript_id_clean"]:
                manuscript_key = (
                    f"{sheet_config.journal_code}::__missing__::{sheet_config.source_file}::"
                    f"{sheet_config.source_sheet}::{row_number}"
                )
                conn.execute(
                    """
                    INSERT INTO etl_warnings
                    (journal_code, source_file, source_sheet, source_row_number, warning_type, message)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        sheet_config.journal_code,
                        sheet_config.source_file,
                        sheet_config.source_sheet,
                        row_number,
                        "missing_manuscript_id",
                        "稿件编号为空，标准主表用来源行生成临时key",
                    ),
                )
                stats["warnings"] += 1
            else:
                manuscript_key = f"{sheet_config.journal_code}:{values['manuscript_id_clean']}"

            candidate = {
                **values,
                "manuscript_key": manuscript_key,
                "journal_code": sheet_config.journal_code,
                "source_file": sheet_config.source_file,
                "source_sheet": sheet_config.source_sheet,
                "source_row_number": row_number,
                "source_priority": sheet_config.source_priority,
                "imported_at": imported_at,
            }
            candidates[manuscript_key].append(candidate)

            for author_config in author_map:
                author_name = normalize_empty(row_value(row, author_config.author_name_col))
                if not author_name:
                    continue
                candidate_authors[manuscript_key].append(
                    {
                        "manuscript_key": manuscript_key,
                        "journal_code": sheet_config.journal_code,
                        "manuscript_id_clean": values["manuscript_id_clean"],
                        "source_file": sheet_config.source_file,
                        "source_sheet": sheet_config.source_sheet,
                        "source_row_number": row_number,
                        "author_order": author_config.author_order,
                        "author_role": author_config.author_role,
                        "author_name": author_name,
                        "h_index": normalize_empty(row_value(row, author_config.h_index_col)),
                        "institution": normalize_empty(row_value(row, author_config.institution_col)),
                        "country": normalize_empty(row_value(row, author_config.country_col)),
                    }
                )

        conn.execute(
            "UPDATE source_sheets SET data_rows_scanned = ?, data_rows_imported = ? WHERE sheet_id = ?",
            (data_rows_scanned, data_rows_imported, sheet_id),
        )
        stats["source_sheets"] += 1
        workbook.close()

    for manuscript_key, rows in candidates.items():
        rows.sort(
            key=lambda item: (
                item["source_priority"],
                0 if item.get("article_title") else 1,
                item["source_file"],
                item["source_sheet"],
                item["source_row_number"],
            )
        )
        chosen = rows[0]
        duplicate_count = len(rows)
        if duplicate_count > 1:
            stats["duplicates"] += duplicate_count - 1
            conn.execute(
                """
                INSERT INTO etl_warnings
                (journal_code, source_file, source_sheet, source_row_number, warning_type, message)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    chosen["journal_code"],
                    chosen["source_file"],
                    chosen["source_sheet"],
                    chosen["source_row_number"],
                    "duplicate_manuscript_id",
                    f"同一期刊同一稿件编号有 {duplicate_count} 条记录；标准主表保留优先级最高记录",
                ),
            )
            stats["warnings"] += 1

        conn.execute(
            """
            INSERT INTO manuscripts
            (manuscript_key, journal_code, manuscript_id_clean, manuscript_id_raw,
             received_date, revised_date, accepted_date, published_date, declined_date,
             article_title, manuscript_category, current_status, round1_reviewing_date,
             source_file, source_sheet, source_row_number, source_priority, duplicate_count, imported_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                manuscript_key,
                chosen["journal_code"],
                chosen["manuscript_id_clean"],
                chosen["manuscript_id_raw"],
                chosen["received_date"],
                chosen["revised_date"],
                chosen["accepted_date"],
                chosen["published_date"],
                chosen["declined_date"],
                chosen["article_title"],
                chosen["manuscript_category"],
                chosen["current_status"],
                chosen["round1_reviewing_date"],
                chosen["source_file"],
                chosen["source_sheet"],
                chosen["source_row_number"],
                chosen["source_priority"],
                duplicate_count,
                chosen["imported_at"],
            ),
        )
        stats["manuscripts"] += 1

        seen_author_keys = set()
        for author in candidate_authors.get(manuscript_key, []):
            author_key = (
                author["author_order"],
                author["author_role"],
                author["author_name"],
                author["institution"],
                author["country"],
            )
            if author_key in seen_author_keys:
                continue
            seen_author_keys.add(author_key)
            conn.execute(
                """
                INSERT INTO manuscript_authors
                (manuscript_key, journal_code, manuscript_id_clean, source_file, source_sheet,
                 source_row_number, author_order, author_role, author_name, h_index, institution, country)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    author["manuscript_key"],
                    author["journal_code"],
                    author["manuscript_id_clean"],
                    author["source_file"],
                    author["source_sheet"],
                    author["source_row_number"],
                    author["author_order"],
                    author["author_role"],
                    author["author_name"],
                    author["h_index"],
                    author["institution"],
                    author["country"],
                ),
            )
            stats["authors"] += 1

    conn.commit()
    summary_path = sqlite_path.with_name("稿件数据_导入摘要.json")
    summary_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")
    conn.close()
    backup_path = publish_database(working_sqlite_path, sqlite_path)
    stats["sqlite_path"] = str(sqlite_path)
    stats["summary_path"] = str(summary_path)
    stats["backup_path"] = str(backup_path) if backup_path else ""
    return stats


def detect_configured_header_row(worksheet: Any, field_map: dict[str, FieldConfig]) -> int:
    expected_labels = {config.source_label for config in field_map.values() if config.source_label}
    expected_cols = {config.column_index for config in field_map.values() if config.column_index}
    best_score = -1
    best_row = 1
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row or 30, 30), values_only=True), start=1
    ):
        score = 0
        for col_index in expected_cols:
            value = clean_text(row_value(row, col_index))
            if value in expected_labels:
                score += 2
            elif value:
                score += 0
        if score > best_score:
            best_score = score
            best_row = row_index
    return best_row


def main() -> None:
    parser = argparse.ArgumentParser(description="Build manuscript SQLite database from journal Excel workbooks.")
    parser.add_argument("--backup-dir", type=Path, default=DEFAULT_BACKUP_DIR)
    parser.add_argument("--config-dir", type=Path, default=DEFAULT_CONFIG_DIR)
    parser.add_argument("--report-dir", type=Path, default=DEFAULT_REPORT_DIR)
    parser.add_argument("--sqlite-path", type=Path, default=DEFAULT_SQLITE_PATH)
    parser.add_argument("--init-config", action="store_true", help="Regenerate editable CSV config from mapping reports.")
    parser.add_argument("--build", action="store_true", help="Build the SQLite database.")
    args = parser.parse_args()

    if args.init_config:
        init_config(args.report_dir, args.config_dir)
        print(f"配置已生成：{args.config_dir}")

    if args.build or not args.init_config:
        stats = run_build(args.backup_dir, args.config_dir, args.sqlite_path)
        print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
